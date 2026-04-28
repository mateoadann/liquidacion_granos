from __future__ import annotations

import io
import os
import re
from datetime import datetime

from flask import Blueprint, current_app, jsonify, request, send_file
from openpyxl import Workbook
from sqlalchemy.exc import IntegrityError

from ..extensions import db
from ..models import LpgDocument, Taxpayer
from ..time_utils import now_cordoba_naive
from ..middleware import require_auth
from ..services import (
    CertificateValidationError,
    decrypt_secret,
    delete_client_certificates,
    encrypt_secret,
    extract_fecha_liquidacion,
    fecha_liquidacion_as_date,
    fecha_liquidacion_expr,
    get_client_certificate_meta,
    is_placeholder_secret,
    is_valid_ambiente,
    is_valid_cuit,
    save_client_certificates,
    validate_certificate_and_key_paths,
)
from ..services.json_v7_exporter import build_json_v7

clients_bp = Blueprint("clients", __name__)


def _has_clave_fiscal(item: Taxpayer) -> bool:
    if not item.clave_fiscal_encrypted:
        return False
    if is_placeholder_secret(item.clave_fiscal_encrypted):
        return False
    try:
        decrypt_secret(item.clave_fiscal_encrypted)
        return True
    except ValueError:
        return False


def _serialize_client(item: Taxpayer) -> dict:
    return {
        "id": item.id,
        "empresa": item.empresa,
        "cuit": item.cuit,
        "cuit_representado": item.cuit_representado,
        "ambiente": item.ambiente,
        "activo": item.activo,
        "playwright_enabled": item.playwright_enabled,
        "has_clave_fiscal": _has_clave_fiscal(item),
        "cert_crt_filename": item.cert_crt_filename,
        "cert_key_filename": item.cert_key_filename,
        "cert_uploaded_at": item.cert_uploaded_at.isoformat()
        if item.cert_uploaded_at
        else None,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }


def _error(message: str, status_code: int = 400):
    return jsonify({"error": message}), status_code


def _parse_bool(value, field_name: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"true", "1", "yes", "si"}:
            return True
        if normalized in {"false", "0", "no"}:
            return False
    raise ValueError(f"{field_name} debe ser booleano.")


def _parse_active_query(value: str | None) -> bool | None:
    if value is None:
        return None
    normalized = value.strip().lower()
    if normalized in {"true", "1", "yes", "si"}:
        return True
    if normalized in {"false", "0", "no"}:
        return False
    raise ValueError("Parametro 'active' invalido. Use true o false.")


def _parse_export_date(value: str | None) -> datetime | None:
    text = (value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"Fecha inválida: '{text}'. Use DD/MM/AAAA o AAAA-MM-DD.")


def _map_codigo_comprobante(doc: LpgDocument) -> str:
    """Map document type/operation code to Holistor comprobante code.

    - AJUSTE documents → "NL"
    - codTipoOperacion 1 → "F1" (compra)
    - codTipoOperacion 2 → "F2" (venta/consignación)
    - Default → "F1"
    """
    if doc.tipo_documento == "AJUSTE":
        return "NL"
    dl = doc.datos_limpios or {}
    raw = doc.raw_data or {}
    data = dl if dl else (raw.get("data", raw) if isinstance(raw, dict) else {})
    cod = data.get("codTipoOperacion") if isinstance(data, dict) else None
    if str(cod) == "2":
        return "F2"
    return "F1"


def _format_fecha_emision(fecha_str: str | None) -> str:
    """Convert 'YYYY-MM-DD' to 'ddmmyyyy'. Returns '' for None/invalid."""
    if not fecha_str:
        return ""
    try:
        dt = datetime.strptime(str(fecha_str).strip(), "%Y-%m-%d")
        return dt.strftime("%d%m%Y")
    except ValueError:
        return ""


def _build_rpa_row(
    doc: LpgDocument, empresa: str, mes: str, anio: str
) -> dict:
    """Build a single row for the RPA/Holistor export XLSX."""
    coe = doc.coe or ""
    tipo_pto_vta = coe[:4] if len(coe) >= 4 else coe
    nro_comprobante = coe[4:] if len(coe) > 4 else ""
    fecha_emision = _format_fecha_emision(extract_fecha_liquidacion(doc))
    return {
        "empresa": empresa,
        "mes": mes,
        "anio": anio,
        "codigo_comprobante": _map_codigo_comprobante(doc),
        "tipo_pto_vta": tipo_pto_vta,
        "nro_comprobante": nro_comprobante,
        "fecha_emision": fecha_emision,
    }


# Column order for the RPA export spreadsheet
_RPA_FIELDNAMES = [
    "empresa",
    "mes",
    "anio",
    "codigo_comprobante",
    "tipo_pto_vta",
    "nro_comprobante",
    "fecha_emision",
]


def _sanitize_filename(text: str) -> str:
    """Sanitize a string for safe use in a filename."""
    safe = re.sub(r"[^\w\-]", "_", text.strip())
    safe = re.sub(r"_+", "_", safe).strip("_")
    return safe or "export"


@clients_bp.get("/clients")
@require_auth
def list_clients():
    try:
        active = _parse_active_query(request.args.get("active"))
    except ValueError as exc:
        return _error(str(exc), 400)

    query = Taxpayer.query.order_by(Taxpayer.id.asc())
    if active is True:
        query = query.filter(Taxpayer.activo.is_(True))
    elif active is False:
        query = query.filter(Taxpayer.activo.is_(False))

    return jsonify([_serialize_client(item) for item in query.all()])


@clients_bp.post("/clients")
@require_auth
def create_client():
    payload = request.get_json(silent=True) or {}

    empresa = str(payload.get("empresa", "")).strip()
    cuit = str(payload.get("cuit", "")).strip()
    cuit_representado = str(payload.get("cuit_representado", "")).strip()
    ambiente = str(payload.get("ambiente", "homologacion")).strip().lower()
    clave_fiscal = payload.get("clave_fiscal")

    if not empresa:
        return _error("empresa es obligatoria.", 400)
    if not is_valid_cuit(cuit):
        return _error("CUIT invalida. Debe tener 11 digitos.", 400)
    if not is_valid_cuit(cuit_representado):
        return _error("cuit_representado invalido. Debe tener 11 digitos.", 400)
    if not is_valid_ambiente(ambiente):
        return _error("ambiente invalido. Valores permitidos: homologacion, produccion.", 400)
    if not isinstance(clave_fiscal, str) or not clave_fiscal.strip():
        return _error("clave_fiscal es obligatoria.", 400)
    if Taxpayer.query.filter_by(cuit=cuit).first():
        return _error("La CUIT ya existe.", 409)

    try:
        encrypted_secret = encrypt_secret(clave_fiscal)
    except ValueError as exc:
        return _error(str(exc), 400)

    item = Taxpayer()
    item.empresa = empresa
    item.cuit = cuit
    item.cuit_representado = cuit_representado
    item.ambiente = ambiente
    item.clave_fiscal_encrypted = encrypted_secret
    item.activo = True
    item.playwright_enabled = True
    db.session.add(item)

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return _error("La CUIT ya existe.", 409)

    return jsonify(_serialize_client(item)), 201


@clients_bp.get("/clients/<int:client_id>")
@require_auth
def get_client(client_id: int):
    item = Taxpayer.query.get_or_404(client_id)
    return jsonify(_serialize_client(item))


@clients_bp.patch("/clients/<int:client_id>")
@require_auth
def update_client(client_id: int):
    item = Taxpayer.query.get_or_404(client_id)
    payload = request.get_json(silent=True) or {}

    if "empresa" in payload:
        empresa = str(payload.get("empresa", "")).strip()
        if not empresa:
            return _error("empresa no puede estar vacia.", 400)
        item.empresa = empresa

    if "cuit" in payload:
        cuit = str(payload.get("cuit", "")).strip()
        if not is_valid_cuit(cuit):
            return _error("CUIT invalida. Debe tener 11 digitos.", 400)
        existing = Taxpayer.query.filter(Taxpayer.cuit == cuit, Taxpayer.id != item.id).first()
        if existing:
            return _error("La CUIT ya existe.", 409)
        item.cuit = cuit

    if "cuit_representado" in payload:
        cuit_representado = str(payload.get("cuit_representado", "")).strip()
        if not is_valid_cuit(cuit_representado):
            return _error("cuit_representado invalido. Debe tener 11 digitos.", 400)
        item.cuit_representado = cuit_representado

    if "ambiente" in payload:
        ambiente = str(payload.get("ambiente", "")).strip().lower()
        if not is_valid_ambiente(ambiente):
            return _error(
                "ambiente invalido. Valores permitidos: homologacion, produccion.", 400
            )
        item.ambiente = ambiente

    if "activo" in payload:
        try:
            item.activo = _parse_bool(payload["activo"], "activo")
        except ValueError as exc:
            return _error(str(exc), 400)

    if "playwright_enabled" in payload:
        try:
            item.playwright_enabled = _parse_bool(
                payload["playwright_enabled"], "playwright_enabled"
            )
        except ValueError as exc:
            return _error(str(exc), 400)

    if "clave_fiscal" in payload:
        clave_fiscal = payload.get("clave_fiscal")
        if not isinstance(clave_fiscal, str) or not clave_fiscal.strip():
            return _error("clave_fiscal no puede estar vacia.", 400)
        try:
            item.clave_fiscal_encrypted = encrypt_secret(clave_fiscal)
        except ValueError as exc:
            return _error(str(exc), 400)

    item.updated_at = now_cordoba_naive()

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return _error("La CUIT ya existe.", 409)

    return jsonify(_serialize_client(item))


@clients_bp.delete("/clients/<int:client_id>")
@require_auth
def delete_client(client_id: int):
    item = Taxpayer.query.get_or_404(client_id)
    item.activo = False
    item.updated_at = now_cordoba_naive()
    db.session.commit()
    return jsonify(_serialize_client(item))


@clients_bp.post("/clients/<int:client_id>/generate-csr")
@require_auth
def generate_client_csr(client_id: int):
    """Genera key + CSR para el cliente usando sus datos (cuit_representado)."""
    from pathlib import Path

    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import rsa
    from cryptography import x509 as x509_mod
    from cryptography.x509.oid import NameOID

    item = Taxpayer.query.get_or_404(client_id)

    payload = request.get_json(silent=True) or {}
    cert_name = (payload.get("nombre_certificado") or "").strip()
    if not cert_name or not re.match(r"^[a-zA-Z0-9_]+$", cert_name):
        return jsonify({"error": "nombre_certificado es requerido (solo alfanumerico y guion bajo)"}), 400

    if not is_valid_cuit(item.cuit_representado):
        return jsonify({"error": "El cliente no tiene un CUIT representado valido"}), 400

    # Generar key RSA 2048
    private_key = rsa.generate_private_key(public_exponent=65537, key_size=2048)

    # Guardar key en filesystem del cliente
    base_path = Path(
        current_app.config.get("CLIENT_CERTIFICATES_BASE_PATH")
        or os.environ.get("CLIENT_CERTIFICATES_BASE_PATH", "/app/certificados_clientes")
    )
    client_dir = base_path / str(item.id)
    client_dir.mkdir(parents=True, exist_ok=True)

    key_path = client_dir / "private.key"
    key_pem = private_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.TraditionalOpenSSL,
        encryption_algorithm=serialization.NoEncryption(),
    )
    key_path.write_bytes(key_pem)
    key_path.chmod(0o600)

    # Actualizar paths en DB
    item.cert_key_path = str(key_path)
    item.cert_key_filename = f"{cert_name}.key"
    item.cert_crt_path = None
    item.cert_crt_filename = None
    item.cert_uploaded_at = None
    item.updated_at = now_cordoba_naive()
    db.session.commit()

    # Generar CSR
    subject = x509_mod.Name([
        x509_mod.NameAttribute(NameOID.COMMON_NAME, cert_name),
        x509_mod.NameAttribute(NameOID.SERIAL_NUMBER, f"CUIT {item.cuit_representado}"),
    ])
    csr = (
        x509_mod.CertificateSigningRequestBuilder()
        .subject_name(subject)
        .sign(private_key, hashes.SHA256())
    )
    csr_pem = csr.public_bytes(serialization.Encoding.PEM)

    return send_file(
        io.BytesIO(csr_pem),
        mimetype="application/pkcs10",
        as_attachment=True,
        download_name=f"{cert_name}.csr",
    )


@clients_bp.post("/clients/<int:client_id>/certificates")
@require_auth
def upload_client_certificates(client_id: int):
    item = Taxpayer.query.get_or_404(client_id)

    cert_file = request.files.get("cert_file")
    key_file = request.files.get("key_file")

    if cert_file is None or key_file is None:
        return _error("Debe enviar cert_file y key_file.", 400)

    try:
        saved_meta = save_client_certificates(item.id, cert_file, key_file)
    except CertificateValidationError as exc:
        return _error(str(exc), 422)
    except ValueError as exc:
        return _error(str(exc), 400)

    item.cert_crt_path = saved_meta["cert_crt_path"]
    item.cert_key_path = saved_meta["cert_key_path"]
    item.cert_crt_filename = saved_meta["cert_crt_filename"]
    item.cert_key_filename = saved_meta["cert_key_filename"]
    item.cert_uploaded_at = now_cordoba_naive()
    item.updated_at = now_cordoba_naive()

    db.session.commit()

    return jsonify(
        {
            "client": _serialize_client(item),
            "certificates": get_client_certificate_meta(item.id),
        }
    )


@clients_bp.get("/clients/<int:client_id>/certificates/meta")
@require_auth
def get_certificates_meta(client_id: int):
    item = Taxpayer.query.get_or_404(client_id)
    storage_meta = get_client_certificate_meta(item.id)

    return jsonify(
        {
            "client_id": item.id,
            "cert_crt_filename": item.cert_crt_filename,
            "cert_key_filename": item.cert_key_filename,
            "cert_uploaded_at": item.cert_uploaded_at.isoformat()
            if item.cert_uploaded_at
            else None,
            "has_certificates": storage_meta["has_certificates"],
            "storage": storage_meta,
        }
    )


@clients_bp.delete("/clients/<int:client_id>/certificates")
@require_auth
def remove_client_certificates(client_id: int):
    item = Taxpayer.query.get_or_404(client_id)
    delete_client_certificates(item.id)

    item.cert_crt_path = None
    item.cert_key_path = None
    item.cert_crt_filename = None
    item.cert_key_filename = None
    item.cert_uploaded_at = None
    item.updated_at = now_cordoba_naive()
    db.session.commit()

    return jsonify({"ok": True, "message": "Certificados eliminados."})


@clients_bp.post("/clients/<int:client_id>/validate-config")
@require_auth
def validate_client_config(client_id: int):
    item = Taxpayer.query.get_or_404(client_id)
    storage_meta = get_client_certificate_meta(item.id)

    has_empresa = bool(item.empresa and item.empresa.strip())
    has_cuit = is_valid_cuit(item.cuit)
    has_cuit_representado = is_valid_cuit(item.cuit_representado)
    has_clave_fiscal = _has_clave_fiscal(item)
    has_certificates = bool(
        item.cert_crt_path
        and item.cert_key_path
        and storage_meta["cert_crt_exists"]
        and storage_meta["cert_key_exists"]
    )

    certificates_valid = False
    if has_certificates:
        try:
            validate_certificate_and_key_paths(item.cert_crt_path, item.cert_key_path)
            certificates_valid = True
        except CertificateValidationError:
            certificates_valid = False

    ready_for_playwright = all(
        [
            has_empresa,
            has_cuit,
            has_cuit_representado,
            has_clave_fiscal,
            has_certificates,
            certificates_valid,
            item.activo,
            item.playwright_enabled,
        ]
    )

    return jsonify(
        {
            "has_empresa": has_empresa,
            "has_cuit": has_cuit,
            "has_cuit_representado": has_cuit_representado,
            "has_clave_fiscal": has_clave_fiscal,
            "has_certificates": has_certificates,
            "certificates_valid": certificates_valid,
            "ready_for_playwright": ready_for_playwright,
        }
    )


@clients_bp.post("/clients/<int:client_id>/test-certificates")
@require_auth
def test_client_certificates(client_id: int):
    """Prueba real de certificados contra ARCA: config local + WSLPG + Constancia."""
    import os
    from pathlib import Path
    from cryptography import x509

    from ..integrations.arca.client import (
        ArcaConstanciaClient,
        ArcaDiscoveryConfig,
        ArcaIntegrationError,
        ArcaWslpgClient,
    )

    item = Taxpayer.query.get_or_404(client_id)

    # --- 1) Config local (reutiliza lógica de validate-config) ---
    storage_meta = get_client_certificate_meta(item.id)
    checks = {
        "has_empresa": bool(item.empresa and item.empresa.strip()),
        "has_cuit": is_valid_cuit(item.cuit),
        "has_cuit_representado": is_valid_cuit(item.cuit_representado),
        "has_clave_fiscal": _has_clave_fiscal(item),
        "has_certificates": bool(
            item.cert_crt_path
            and item.cert_key_path
            and storage_meta["cert_crt_exists"]
            and storage_meta["cert_key_exists"]
        ),
        "certificates_valid": False,
    }

    if checks["has_certificates"]:
        try:
            validate_certificate_and_key_paths(item.cert_crt_path, item.cert_key_path)
            checks["certificates_valid"] = True
        except CertificateValidationError:
            pass

    config_ok = all(checks.values())

    # --- 2) Info del certificado ---
    cert_info = None
    if checks["has_certificates"]:
        try:
            cert_data = open(item.cert_crt_path, "rb").read()
            cert = x509.load_pem_x509_certificate(cert_data)
            now = datetime.utcnow()
            cert_info = {
                "subject": cert.subject.rfc4514_string(),
                "issuer": cert.issuer.rfc4514_string(),
                "not_before": cert.not_valid_before_utc.strftime("%Y-%m-%d"),
                "not_after": cert.not_valid_after_utc.strftime("%Y-%m-%d"),
                "expired": now > cert.not_valid_after_utc.replace(tzinfo=None),
            }
        except Exception:
            pass

    # --- 3) Test WSLPG ---
    wslpg_result = {"ok": False, "message": "No ejecutado"}
    if config_ok:
        try:
            config = ArcaDiscoveryConfig.from_env()
            config.environment = item.ambiente or config.environment
            config.cuit_representada = item.cuit_representado
            config.cert_path = item.cert_crt_path
            config.key_path = item.cert_key_path
            ta_base = config.ta_path or os.getenv("ARCA_TA_PATH") or "/tmp/ta"
            config.ta_path = str(Path(ta_base) / f"taxpayer_{item.id}")
            ws = ArcaWslpgClient(config=config)
            ws.call_dummy()
            wslpg_result = {"ok": True, "message": "Conexión exitosa a WSLPG"}
        except (ArcaIntegrationError, Exception) as exc:
            wslpg_result = {"ok": False, "message": str(exc)}

    # --- 4) Test Constancia ---
    constancia_result = {"ok": False, "message": "No ejecutado"}
    if config_ok:
        try:
            config = ArcaDiscoveryConfig.from_env()
            config.environment = item.ambiente or config.environment
            config.cuit_representada = item.cuit_representado
            config.cert_path = item.cert_crt_path
            config.key_path = item.cert_key_path
            ta_base = config.ta_path or os.getenv("ARCA_TA_PATH") or "/tmp/ta"
            config.ta_path = str(Path(ta_base) / f"taxpayer_{item.id}_constancia")
            client = ArcaConstanciaClient(config=config)
            info = client.extract_persona_info(item.cuit_representado)
            constancia_result = {
                "ok": True,
                "message": "Conexión exitosa a Padrón",
                "razonSocial": info.get("razonSocial", ""),
            }
        except (ArcaIntegrationError, Exception) as exc:
            constancia_result = {"ok": False, "message": str(exc)}

    return jsonify({
        "config": {"ok": config_ok, "checks": checks},
        "wslpg": wslpg_result,
        "constancia": constancia_result,
        "certificate_info": cert_info,
    })


@clients_bp.get("/clients/<int:client_id>/coes/export")
@require_auth
def export_client_coes(client_id: int):
    client = Taxpayer.query.get_or_404(client_id)

    # --- Parse and validate date range ---
    try:
        fecha_desde = _parse_export_date(request.args.get("fecha_desde"))
        fecha_hasta = _parse_export_date(request.args.get("fecha_hasta"))
    except ValueError as exc:
        return _error(str(exc), 400)

    if fecha_desde is None or fecha_hasta is None:
        return _error("fecha_desde and fecha_hasta are required", 400)

    # Same calendar-month validation
    if (
        fecha_desde.month != fecha_hasta.month
        or fecha_desde.year != fecha_hasta.year
    ):
        return _error(
            "fecha_desde and fecha_hasta must be in the same calendar month",
            400,
        )

    mes = str(fecha_desde.month)
    anio = str(fecha_desde.year)

    # --- Query documents ---
    query = LpgDocument.query.filter(LpgDocument.taxpayer_id == client.id)

    fecha_liq_expr = fecha_liquidacion_expr()
    fecha_liq_date = fecha_liquidacion_as_date(fecha_liq_expr)
    query = query.filter(fecha_liq_date >= fecha_desde.date())
    query = query.filter(fecha_liq_date <= fecha_hasta.date())

    documents = query.order_by(
        fecha_liq_date.asc(), LpgDocument.id.asc()
    ).all()

    # --- Build rows ---
    rows = [
        _build_rpa_row(doc, client.empresa or "", mes, anio)
        for doc in documents
    ]

    # --- Generate XLSX with openpyxl ---
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "COEs"

    # Header
    sheet.append(_RPA_FIELDNAMES)

    # Text-column indices (1-based): tipo_pto_vta=5, nro_comprobante=6, fecha_emision=7
    # These columns contain numeric-looking strings that must remain as text
    # (leading zeros in nro_comprobante, ddmmyyyy in fecha_emision, etc.).
    _TEXT_COLS = {5, 6, 7}

    for row_data in rows:
        row_idx = sheet.max_row + 1
        for col_idx, key in enumerate(_RPA_FIELDNAMES, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if col_idx in _TEXT_COLS:
                # Write as explicit text: set format BEFORE value so openpyxl
                # does not infer a numeric type from the string content.
                cell.number_format = "@"
                cell.value = str(row_data[key])
            else:
                cell.value = row_data[key]

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    empresa_safe = _sanitize_filename(client.empresa or f"cliente_{client.id}")
    filename = f"{empresa_safe}_{mes}_{anio}_coes.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@clients_bp.get("/clients/<int:client_id>/export/json-v7")
@require_auth
def export_json_v7(client_id):
    taxpayer = Taxpayer.query.get(client_id)
    if taxpayer is None:
        return _error("Cliente no encontrado.", 404)

    # --- Validate required params: mes, anio ---
    raw_mes = request.args.get("mes")
    raw_anio = request.args.get("anio")
    if raw_mes is None or raw_anio is None:
        return _error("mes y anio son obligatorios.", 400)

    try:
        mes = int(raw_mes)
        anio = int(raw_anio)
    except (ValueError, TypeError):
        return _error("mes y anio deben ser enteros.", 400)

    if mes < 1 or mes > 12:
        return _error("mes debe estar entre 1 y 12.", 400)

    # --- Parse optional date range ---
    try:
        fecha_desde = _parse_export_date(request.args.get("fecha_desde"))
        fecha_hasta = _parse_export_date(request.args.get("fecha_hasta"))
    except ValueError as exc:
        return _error(str(exc), 400)

    # --- Query documents ---
    query = LpgDocument.query.filter(LpgDocument.taxpayer_id == taxpayer.id)
    query = query.filter(LpgDocument.datos_limpios.isnot(None))

    fecha_liq_expr = fecha_liquidacion_expr()
    fecha_liq_date = fecha_liquidacion_as_date(fecha_liq_expr)

    if fecha_desde is not None:
        query = query.filter(fecha_liq_date >= fecha_desde.date())
    if fecha_hasta is not None:
        query = query.filter(fecha_liq_date <= fecha_hasta.date())

    documents = query.order_by(
        fecha_liq_date.asc(), LpgDocument.id.asc()
    ).all()

    # --- Build JSON v7 ---
    result = build_json_v7(documents, taxpayer, mes, anio)

    response = jsonify(result)
    cuit = taxpayer.cuit_representado or "00000000000"
    filename = f"liquidaciones_v7_{cuit}_{mes}_{anio}.json"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
