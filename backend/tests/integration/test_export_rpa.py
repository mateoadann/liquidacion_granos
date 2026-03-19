from __future__ import annotations

import io

from openpyxl import load_workbook

from app.extensions import db
from app.models import LpgDocument, Taxpayer


def _create_taxpayer(*, cuit: str = "20304050607", empresa: str = "Acopio SA") -> Taxpayer:
    item = Taxpayer()
    item.cuit = cuit
    item.empresa = empresa
    item.cuit_representado = cuit
    item.clave_fiscal_encrypted = "test"
    item.activo = True
    db.session.add(item)
    db.session.commit()
    return item


def _create_coe(
    *,
    taxpayer_id: int,
    coe: str = "330230787118",
    tipo_documento: str = "LPG",
    datos_limpios: dict | None = None,
    raw_data: dict | None = None,
) -> LpgDocument:
    doc = LpgDocument()
    doc.taxpayer_id = taxpayer_id
    doc.coe = coe
    doc.tipo_documento = tipo_documento
    doc.datos_limpios = datos_limpios
    doc.raw_data = raw_data
    db.session.add(doc)
    db.session.commit()
    return doc


def _export_url(client_id: int) -> str:
    return f"/api/clients/{client_id}/coes/export"


def _read_xlsx(response_data: bytes) -> list[list]:
    """Parse XLSX response bytes and return rows as lists (header + data)."""
    wb = load_workbook(filename=io.BytesIO(response_data))
    ws = wb.active
    return [[cell.value for cell in row] for row in ws.iter_rows()]


# ─── Validation tests ────────────────────────────────────────────────


def test_export_missing_fecha_desde_returns_400(client, auth_headers):
    taxpayer = _create_taxpayer()

    response = client.get(
        _export_url(taxpayer.id),
        query_string={"fecha_hasta": "2026-02-28"},
        headers=auth_headers,
    )

    assert response.status_code == 400
    body = response.get_json()
    assert "fecha_desde" in body["error"].lower() or "required" in body["error"].lower()


def test_export_missing_fecha_hasta_returns_400(client, auth_headers):
    taxpayer = _create_taxpayer()

    response = client.get(
        _export_url(taxpayer.id),
        query_string={"fecha_desde": "2026-02-01"},
        headers=auth_headers,
    )

    assert response.status_code == 400
    body = response.get_json()
    assert "fecha_hasta" in body["error"].lower() or "required" in body["error"].lower()


def test_export_dates_different_months_returns_400(client, auth_headers):
    taxpayer = _create_taxpayer()

    response = client.get(
        _export_url(taxpayer.id),
        query_string={
            "fecha_desde": "2026-02-01",
            "fecha_hasta": "2026-03-15",
        },
        headers=auth_headers,
    )

    assert response.status_code == 400
    body = response.get_json()
    assert "same" in body["error"].lower() or "month" in body["error"].lower()


# ─── Successful export tests ─────────────────────────────────────────


def test_export_valid_dates_returns_xlsx(client, auth_headers):
    taxpayer = _create_taxpayer()
    _create_coe(
        taxpayer_id=taxpayer.id,
        coe="330230787118",
        datos_limpios={
            "codTipoOperacion": 1,
            "fechaLiquidacion": "2026-02-15",
        },
    )

    response = client.get(
        _export_url(taxpayer.id),
        query_string={
            "fecha_desde": "2026-02-01",
            "fecha_hasta": "2026-02-28",
        },
        headers=auth_headers,
    )

    assert response.status_code == 200
    assert (
        "spreadsheetml" in response.content_type
        or "application/vnd.openxmlformats" in response.content_type
    )


def test_export_xlsx_has_correct_columns(client, auth_headers):
    taxpayer = _create_taxpayer()
    _create_coe(
        taxpayer_id=taxpayer.id,
        coe="330230787118",
        datos_limpios={
            "codTipoOperacion": 2,
            "fechaLiquidacion": "2026-02-10",
        },
    )

    response = client.get(
        _export_url(taxpayer.id),
        query_string={
            "fecha_desde": "2026-02-01",
            "fecha_hasta": "2026-02-28",
        },
        headers=auth_headers,
    )

    assert response.status_code == 200
    rows = _read_xlsx(response.data)

    expected_headers = [
        "empresa",
        "mes",
        "anio",
        "codigo_comprobante",
        "tipo_pto_vta",
        "nro_comprobante",
        "fecha_emision",
    ]
    assert rows[0] == expected_headers
    assert len(rows) == 2  # header + 1 data row

    data_row = rows[1]
    assert data_row[0] == "Acopio SA"  # empresa
    assert data_row[3] == "F2"  # codigo_comprobante (codTipoOperacion=2)
    assert data_row[4] == "3302"  # tipo_pto_vta
    assert data_row[6] == "10022026"  # fecha_emision


def test_export_nro_comprobante_preserves_leading_zeros(client, auth_headers):
    """Ensure nro_comprobante is stored as text, preserving leading zeros."""
    taxpayer = _create_taxpayer()
    # COE with leading zero in the nro_comprobante part: "0055" + "00787118"
    _create_coe(
        taxpayer_id=taxpayer.id,
        coe="005500787118",
        datos_limpios={
            "codTipoOperacion": 1,
            "fechaLiquidacion": "2026-02-20",
        },
    )

    response = client.get(
        _export_url(taxpayer.id),
        query_string={
            "fecha_desde": "2026-02-01",
            "fecha_hasta": "2026-02-28",
        },
        headers=auth_headers,
    )

    assert response.status_code == 200
    rows = _read_xlsx(response.data)
    data_row = rows[1]

    # tipo_pto_vta and nro_comprobante must be text, not stripped of zeros
    assert data_row[4] == "0055"
    assert data_row[5] == "00787118"
